from rest_framework.pagination import PageNumberPagination
from rest_framework import status
from datetime import datetime
from openpyxl import load_workbook
from openpyxl import Workbook
import pandas as pd
from Employee.models import Employee
from Salary.models import Salary
from Timekeeping.models import Timekeeping, TypeCheckChoicesEnum
from fcm_django.models import FCMDevice
import gspread
from google.oauth2 import service_account
from googleapiclient.discovery import build
from django.http import HttpResponse
from django.conf import settings

import os.path

class LargeResultsSetPagination(PageNumberPagination):
    page_size = 6
    page_size_query_param = 'page_size'
    max_page_size = 100


def format_respone(success, status, message, data):
    return {
        "success" : success,
        "status" : status,
        "message" : message,
        "data" : data, 
    }

def get_paginated_response(success, statusCode, message, totalItems, nextPage, previousPage, currentPage, totalPages, limit, data):
    return{
        "success": success,
        "statusCode": statusCode,
        "message": message,
        "totalItems": totalItems,
        "nextPage": nextPage,
        "previousPage": previousPage,
        "currentPage": currentPage,
        "totalPages": totalPages,
        "limit": limit,
        "data": data,
    }
    
def create_paginated_response(serializer, queryset, paginator, success=True, status_code=status.HTTP_200_OK):
    
    currentpage = paginator.page.number if paginator.page else None
    message = f"Get Page {currentpage} successfully"
    
    return get_paginated_response(
        success=success,
        statusCode=status_code,
        message=message,
        totalItems=queryset.count(),
        nextPage=paginator.page.next_page_number() if paginator.page and paginator.page.has_next() else None,
        previousPage=paginator.page.previous_page_number() if paginator.page and paginator.page.has_previous() else None,
        currentPage=currentpage,
        totalPages=paginator.page.paginator.num_pages if paginator.page else None,
        limit=paginator.page.paginator.per_page if paginator.page else None,
        data=serializer.data,
    )
    
def format_date(date):
    return date.strftime("%d/%m/%Y")

def format_time(time_obj):
    return time_obj.strftime("%H:%M")

# def update_excel_file(employee, check_type):
#     file_path = 'check_in_out_records.xlsx'
    
#     if os.path.exists(file_path):
#         # Tệp đã tồn tại, mở nó để cập nhật

#         wb = load_workbook(file_path)
#         ws = wb.active
#     else:
#         # Tạo một workbook mới nếu tệp không tồn tại

#         wb = Workbook()
#         ws = wb.active
#         ws.append(['Employee', 'Check Type', 'Date', 'Time'])

#     # Thêm dữ liệu vào worksheet
#     ws.append([employee.first_name + " " + employee.last_name , check_type, datetime.now().date(), datetime.now().time()])
#     # Lưu workbook vào tệp Excel
#     wb.save(file_path)

# def update_excel_file_salary(employee, total_work_days, total_leave_days, total_salary, month, year):
#     file_path = 'salary.xlsx'
    
#     if os.path.exists(file_path):
#         # Tệp đã tồn tại, mở nó để cập nhật

#         wb = load_workbook(file_path)
#         ws = wb.active
#     else:
#         # Tạo một workbook mới nếu tệp không tồn tại

#         wb = Workbook()
#         ws = wb.active
#         ws.append(['Employee', 'Month', 'Year', 'Total Work Days', 'Total Leave Days', 'Total Salary'])

#     # Thêm dữ liệu vào worksheet
#     ws.append([employee.first_name + " " + employee.last_name , month, year, total_work_days, total_leave_days, int(total_salary)])
#     # Lưu workbook vào tệp Excel
#     wb.save(file_path)
    

def export_salary_excel(queryset):
    # Define the default directory to save the Excel file
    default_directory = settings.MEDIA_ROOT

    # Create Excel workbook and sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Salaries"

    # Write headers
    headers = ["Employee", "Month", "Year", "Total Work Days", "Total Leave Days", "Total Salary"]
    ws.append(headers)

    # Write data
    for salary in queryset:
        ws.append([salary['employee_fullname'], salary['salary_month'], salary['salary_year'], salary['total_workdays'], salary["total_leave_days"], salary["total_salary"]])

    # Define the file path
    excel_file_path = os.path.join(default_directory,'excels','salaries.xlsx')

    # Save the workbook to the defined file path
    wb.save(excel_file_path)

    return excel_file_path

def export_time_keeping_excel(queryset):
    # Define the default directory to save the Excel file
    default_directory = settings.MEDIA_ROOT

    # Create Excel workbook and sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Timekeeping"

    # Write headers
    headers = ['ID Employee', 'Fullname', 'Date', 'Weekday', 'Check In', 'Check Out']
    ws.append(headers)

    # Write data
    added_records = set()
    
    for timekeeping in queryset:
    # Tạo một đối tượng ngày từ ngày kiểm tra
        check_date = timekeeping['date_check']
        # print(check_date)
    
    # Xác định ngày trong tuần (0: Thứ 2, 1: Thứ 3, ..., 6: Chủ nhật)
        weekday_mapping = {
        0: "Thứ 2",
        1: "Thứ 3",
        2: "Thứ 4",
        3: "Thứ 5",
        4: "Thứ 6",
        5: "Thứ 7",
        6: "Chủ nhật"
        }

        weekday = check_date.weekday()
        weekday_string = weekday_mapping[weekday]
    
        # Kiểm tra xem bản ghi này đã tồn tại trong danh sách data chưa
        record_key = (timekeeping['employee_id'], check_date)
        if record_key in added_records:
              continue
    
        # Nếu chưa, thêm vào danh sách data và cập nhật tập hợp các bản ghi đã thêm
        added_records.add(record_key)
    
        # Lấy thông tin check in
        obj_check_in = Timekeeping.objects.filter(employee=timekeeping['employee_id'], date=check_date, check_type=TypeCheckChoicesEnum.CHECKIN).values('check_time').first()
        check_in_time = obj_check_in['check_time'] if obj_check_in else None
    
        # Lấy thông tin check out
        obj_check_out = Timekeeping.objects.filter(employee=timekeeping['employee_id'], date=check_date, check_type=TypeCheckChoicesEnum.CHECKOUT).values('check_time').first()
        check_out_time = obj_check_out['check_time'] if obj_check_out else None
    
        # Thêm dữ liệu vào danh sách data
        ws.append([
            timekeeping['employee_id'],
            timekeeping['employee'],
            format_date(check_date),  # Định dạng ngày
            weekday_string,  # Thêm ngày trong tuần
            format_time(check_in_time) if check_in_time else '',  # Định dạng giờ check in
            format_time(check_out_time) if check_out_time else '',  # Định dạng giờ check out
        ])

    # Define the file path
    excel_file_path = os.path.join(default_directory,'excels','timekeeping.xlsx')

    # Save the workbook to the defined file path
    wb.save(excel_file_path)

    return excel_file_path

def update_google_sheet_salary(queryset):
    # Đường dẫn tới tệp JSON khóa dịch vụ
    credentials_file = 'exnodes-8257f-130c271f8282.json'

    # Phạm vi (scope) của dịch vụ Google Sheets API
    SCOPES = ['https://www.googleapis.com/auth/spreadsheets']

    # Xác thực với dịch vụ Google Sheets API
    credentials = service_account.Credentials.from_service_account_file(credentials_file, scopes=SCOPES)
    service = build('sheets', 'v4', credentials=credentials)

    # ID của Google Sheets
    spreadsheet_id = '1NqfqLSMXjt32H4UJaDwR7NOarUQUbx0Cu_rcDAa9E_0'
    
    clear_range = 'Salary!A1:F'

    # Xóa dữ liệu trong phạm vi đã chỉ định
    response = service.spreadsheets().values().clear(
        spreadsheetId=spreadsheet_id,
        range=clear_range,
    ).execute()
    
    
    data = [
        ["Employee", "Month", "Year", "Total Work Days", "Total Leave Days", "Total Salary"]
    ]
    
    for salary in queryset:
        data.append([salary['employee_fullname'], salary['salary_month'], salary['salary_year'], salary['total_workdays'], salary["total_leave_days"], salary["total_salary"]])

    # data.append([])
    
    num_rows = len(data)
    print(num_rows)

    # Range của dữ liệu (ví dụ: Sheet1!A1:F6)
    range_ = 'Salary!A1:F{}'.format(num_rows)
    
    request = {
    'valueInputOption': 'RAW',
    'data': [
        {
            'range': range_,
            'values': data,
            'majorDimension': 'ROWS'
        }
    ]
}


    # Ghi dữ liệu vào Google Sheets
    response = service.spreadsheets().values().batchUpdate(
        spreadsheetId=spreadsheet_id,
        body=request,
    ).execute()
    
    spreadsheet_url = response.get('spreadsheetUrl', '')

    return spreadsheet_url
    
def update_google_sheet_timekeeping(queryset):
    # Đường dẫn tới tệp JSON khóa dịch vụ
    credentials_file = 'exnodes-8257f-130c271f8282.json'

    # Phạm vi (scope) của dịch vụ Google Sheets API
    SCOPES = ['https://www.googleapis.com/auth/spreadsheets']

    # Xác thực với dịch vụ Google Sheets API
    credentials = service_account.Credentials.from_service_account_file(credentials_file, scopes=SCOPES)
    service = build('sheets', 'v4', credentials=credentials)

    # ID của Google Sheets
    spreadsheet_id = '1NqfqLSMXjt32H4UJaDwR7NOarUQUbx0Cu_rcDAa9E_0'
    
    clear_range = 'Timekeeping!A1:F'

    # Xóa dữ liệu trong phạm vi đã chỉ định
    service.spreadsheets().values().clear(
        spreadsheetId=spreadsheet_id,
        range=clear_range,
    ).execute()
    
    
    data = [
        ['ID Employee', 'Fullname', 'Date', 'Weekday', 'Check In', 'Check Out']
    ]
    
    added_records = set()
    
    for timekeeping in queryset:
    # Tạo một đối tượng ngày từ ngày kiểm tra
        check_date = timekeeping['date_check']
        # print(check_date)
    
    # Xác định ngày trong tuần (0: Thứ 2, 1: Thứ 3, ..., 6: Chủ nhật)
        weekday_mapping = {
        0: "Thứ 2",
        1: "Thứ 3",
        2: "Thứ 4",
        3: "Thứ 5",
        4: "Thứ 6",
        5: "Thứ 7",
        6: "Chủ nhật"
        }

        weekday = check_date.weekday()
        weekday_string = weekday_mapping[weekday]
    
        # Kiểm tra xem bản ghi này đã tồn tại trong danh sách data chưa
        record_key = (timekeeping['employee_id'], check_date)
        if record_key in added_records:
              continue
    
        # Nếu chưa, thêm vào danh sách data và cập nhật tập hợp các bản ghi đã thêm
        added_records.add(record_key)
    
        # Lấy thông tin check in
        obj_check_in = Timekeeping.objects.filter(employee=timekeeping['employee_id'], date=check_date, check_type=TypeCheckChoicesEnum.CHECKIN).values('check_time').first()
        check_in_time = obj_check_in['check_time'] if obj_check_in else None
    
        # Lấy thông tin check out
        obj_check_out = Timekeeping.objects.filter(employee=timekeeping['employee_id'], date=check_date, check_type=TypeCheckChoicesEnum.CHECKOUT).values('check_time').first()
        check_out_time = obj_check_out['check_time'] if obj_check_out else None
    
        # Thêm dữ liệu vào danh sách data
        data.append([
            timekeeping['employee_id'],
            timekeeping['employee'],
            format_date(check_date),  # Định dạng ngày
            weekday_string,  # Thêm ngày trong tuần
            format_time(check_in_time) if check_in_time else '',  # Định dạng giờ check in
            format_time(check_out_time) if check_out_time else '',  # Định dạng giờ check out
        ])

    # data.append([])   
    
    num_rows = len(data)
    # print(num_rows)

    # Range của dữ liệu (ví dụ: Sheet1!A1:F6)
    range_ = 'Timekeeping!A1:F{}'.format(num_rows)
    
    request = {
    'valueInputOption': 'RAW',
    'data': [
        {
            'range': range_,
            'values': data,
            'majorDimension': 'ROWS'
        }
    ]
}


    # Ghi dữ liệu vào Google Sheets
    service.spreadsheets().values().batchUpdate(
        spreadsheetId=spreadsheet_id,
        body=request,
    ).execute()

# def get_data_from_db(employee, month, year, total_work_days, total_leave_days, total_salary):
#     # Lấy dữ liệu từ cơ sở dữ liệu

#     # Chuyển đổi dữ liệu thành danh sách
#     data = [
#         ["Employee", "Month", "Year", "Total Work Days", "Total Leave Days", "Total Salary"]
#     ]
    
#     data.append([employee.last_name, month, year, total_work_days, total_leave_days, int(total_salary)])

#     return data
# def get_admin_devices():
#     # Lấy tất cả các tài khoản admin
#     admins = Employee.objects.filter(is_admin=True)
#     # Lấy tất cả các thiết bị của các tài khoản admin
#     admin_devices = FCMDevice.objects.filter(user__in=admins)
#     print(admins)
#     print(admin_devices)
#     return admin_devices

# def send_push_notification_to_admins(title, body):
#     admin_devices = get_admin_devices()
#     for device in admin_devices:
#         device.send_message(title=title, body=body)